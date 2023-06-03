# -*- coding: utf-8 -*-
"""
Class definition of YOLO_v3 style detection model on image and video
"""
import colorsys
from timeit import default_timer as timer

import numpy as np
from keras import backend as K
from keras.models import load_model
from keras.layers import Input
from PIL import Image, ImageFont, ImageDraw
import openpyxl

from parallel_line import get_right_line,get_left_line

from yolo3.model import yolo_eval, yolo_body, tiny_yolo_body
from yolo3.utils import letterbox_image
import os
from keras.utils import multi_gpu_model
from sort import *
from nms import non_max_suppression_slow
import cv2


os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
import tensorflow as tf
import logging
tf.get_logger().setLevel(logging.ERROR)

import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)

numberOfLines = 1
btn_down = False
lines = []

class YOLO(object):
    _defaults = {
        "model_path": 'model_data/yolo.h5',
        "anchors_path": 'model_data/yolo_anchors.txt',
        "classes_path": 'model_data/coco_classes.txt',
        "score" : 0.3,
        "iou" : 0.45,
        "model_image_size" : (416, 416),
        "gpu_num" : 0,
    }

    @classmethod
    def get_defaults(cls, n):
        if n in cls._defaults:
            return cls._defaults[n]
        else:
            return "Unrecognized attribute name '" + n + "'"

    def __init__(self, **kwargs):
        self.__dict__.update(self._defaults) # set up default values
        self.__dict__.update(kwargs) # and update with user overrides
        self.class_names = self._get_class()
        self.anchors = self._get_anchors()
        self.sess = K.get_session()
        self.boxes, self.scores, self.classes = self.generate()

        # self.numberOfLines = 1
        # self.btn_down = False
        # self.lines = []


        self.counter = 0
        self.left_counter=0
        self.right_counter=0
        self.leftToright_counter=0
        self.rightToleft_counter=0
        self.left=False
        self.right=False
        self.mid=False

        self.tracker = Sort()
        self.memory = {}
        # self.line = [(43, 443), (550, 655)]
        # self.line = [(247, 1), (430, 475)]
        # self.line = [(180, 100), (455, 330)]
        # self.line = [(180, 100), (455, 330)]
        self.line_left = [(111, 134), (311, 432)]
        self.line_right = [(278, 100), (500, 266)]


        np.random.seed(42)
        self.COLORS = np.random.randint(0, 255, size=(200, 3),
                                   dtype="uint8")

    def _get_class(self):
        classes_path = os.path.expanduser(self.classes_path)
        with open(classes_path) as f:
            class_names = f.readlines()
        class_names = [c.strip() for c in class_names]
        return class_names

    def _get_anchors(self):
        anchors_path = os.path.expanduser(self.anchors_path)
        with open(anchors_path) as f:
            anchors = f.readline()
        anchors = [float(x) for x in anchors.split(',')]
        return np.array(anchors).reshape(-1, 2)

    def generate(self):
        model_path = os.path.expanduser(self.model_path)
        assert model_path.endswith('.h5'), 'Keras model or weights must be a .h5 file.'

        # Load model, or construct model and load weights.
        num_anchors = len(self.anchors)
        num_classes = len(self.class_names)
        is_tiny_version = num_anchors==6 # default setting
        try:
            self.yolo_model = load_model(model_path, compile=False)
        except:
            self.yolo_model = tiny_yolo_body(Input(shape=(None,None,3)), num_anchors//2, num_classes) \
                if is_tiny_version else yolo_body(Input(shape=(None,None,3)), num_anchors//3, num_classes)
            self.yolo_model.load_weights(self.model_path) # make sure model, anchors and classes match
        else:
            assert self.yolo_model.layers[-1].output_shape[-1] == \
                num_anchors/len(self.yolo_model.output) * (num_classes + 5), \
                'Mismatch between model and given anchor and class sizes'

        print('{} model, anchors, and classes loaded.'.format(model_path))

        # Generate colors for drawing bounding boxes.
        hsv_tuples = [(x / len(self.class_names), 1., 1.)
                      for x in range(len(self.class_names))]
        self.colors = list(map(lambda x: colorsys.hsv_to_rgb(*x), hsv_tuples))
        self.colors = list(
            map(lambda x: (int(x[0] * 255), int(x[1] * 255), int(x[2] * 255)),
                self.colors))
        np.random.seed(10101)  # Fixed seed for consistent colors across runs.
        np.random.shuffle(self.colors)  # Shuffle colors to decorrelate adjacent classes.
        np.random.seed(None)  # Reset seed to default.

        # Generate output tensor targets for filtered bounding boxes.
        self.input_image_shape = K.placeholder(shape=(2, ))
        if self.gpu_num>=2:
            self.yolo_model = multi_gpu_model(self.yolo_model, gpus=self.gpu_num)
        boxes, scores, classes = yolo_eval(self.yolo_model.output, self.anchors,
                len(self.class_names), self.input_image_shape,
                score_threshold=self.score, iou_threshold=self.iou)
        return boxes, scores, classes

    def ccw(self,A, B, C):
        return (C[1] - A[1]) * (B[0] - A[0]) > (B[1] - A[1]) * (C[0] - A[0])

    def intersect(self,A, B, C, D):
        return self.ccw(A, C, D) != self.ccw(B, C, D) and self.ccw(A, B, C) != self.ccw(A, B, D)

    # from __future__ import division

    def lineFunc(self,p1, p2):
        A = (p1[1] - p2[1])
        B = (p2[0] - p1[0])
        C = (p1[0] * p2[1] - p2[0] * p1[1])
        return A, B, -C

    def intersection(self,L1, L2):
        D = L1[0] * L2[1] - L1[1] * L2[0]
        Dx = L1[2] * L2[1] - L1[1] * L2[2]
        Dy = L1[0] * L2[2] - L1[2] * L2[0]
        if D != 0:
            x = Dx / D
            y = Dy / D
            return x, y
        else:
            return False

        
    def detect_image(self, image,lines_mid):

        # print('lines mid',lines_mid)

        line_right=get_right_line(lines_mid,60)
        line_left=get_left_line(lines_mid,60)

        start = timer()

        if self.model_image_size != (None, None):
            assert self.model_image_size[0]%32 == 0, 'Multiples of 32 required'
            assert self.model_image_size[1]%32 == 0, 'Multiples of 32 required'
            boxed_image = letterbox_image(image, tuple(reversed(self.model_image_size)))
        else:
            new_image_size = (image.width - (image.width % 32),
                              image.height - (image.height % 32))
            boxed_image = letterbox_image(image, new_image_size)
        image_data = np.array(boxed_image, dtype='float32')

        # print(image_data.shape)
        image_data /= 255.
        image_data = np.expand_dims(image_data, 0)  # Add batch dimension.

        out_boxes, out_scores, out_classes = self.sess.run(
            [self.boxes, self.scores, self.classes],
            feed_dict={
                self.yolo_model.input: image_data,
                self.input_image_shape: [image.size[1], image.size[0]],
                K.learning_phase(): 0
            })

        # print('Found {} boxes for {}'.format(len(out_boxes), 'img'))

        # font = ImageFont.truetype(font='font/FiraMono-Medium.otf',
        #             size=np.floor(3e-2 * image.size[1] + 0.5).astype('int32'))
        # thickness = (image.size[0] + image.size[1]) // 300

        dets = []

        font = ImageFont.truetype(font='font/FiraMono-Medium.otf',
                                  size=np.floor(3e-2 * image.size[1] + 0.5).astype('int32'))


        thickness = (image.size[0] + image.size[1]) // 300

        out_boxes = non_max_suppression_slow(out_boxes, 0.3)
        # print(out_boxes)


        for i in range(len(out_boxes)):
        # for i, c in reversed(list(enumerate(out_classes))):

            # print("score",out_scores[i])
            if out_scores[i]>0.2:
                (x1, y1) = (out_boxes[i][0], out_boxes[i][1])
                (x2, y2) = (out_boxes[i][2], out_boxes[i][3])
                # dets.append([ y1, x1, y2,x2, out_scores[i]])
                dets.append([y1, x1, y2, x2, out_classes[i]])

        np.set_printoptions(formatter={'float': lambda x: "{0:0.3f}".format(x)})
        dets = np.asarray(dets)
        # print('sdet', dets)
        tracks = self.tracker.update(dets)

        boxes = []
        indexIDs = []
        c = []
        previous = self.memory.copy()
        self.memory = {}
        # print("previous",previous)
        # print('track',tracks)
        for track in tracks:
            boxes.append([track[0], track[1], track[2], track[3]])
            indexIDs.append(int(track[4]))
            self.memory[indexIDs[-1]] = boxes[-1]

        # print('len box', len(boxes))

        if len(boxes) > 0:
            i = int(0)
            for box in boxes:
                # print('box',box)
                # # extract the bounding box coordinates
                # (x, y) = (int(box[0]), int(box[1]))
                # (w, h) = (int(box[2]), int(box[3]))
                # print('xywh',x,y,w,h)
                # # draw a bounding box rectangle and label on the image
                # # color = [int(c) for c in COLORS[classIDs[i]]]
                # # cv2.rectangle(frame, (x, y), (x + w, y + h), color, 2)
                # frame=np.asarray(image)
                # color = [int(c) for c in self.COLORS[indexIDs[i] % len(self.COLORS)]]
                #
                # cv2.rectangle(frame, (int(box[0]), int(box[1])), (int(box[2]), int(box[3])), (255,0,0), 2)
                # # cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
                #
                # # if indexIDs[i] in previous:
                # #     previous_box = previous[indexIDs[i]]
                # #     (x2, y2) = (int(previous_box[0]), int(previous_box[1]))
                # #     (w2, h2) = (int(previous_box[2]), int(previous_box[3]))
                # #     p0 = (int(x + (w - x) / 2), int(y + (h - y) / 2))
                # #     p1 = (int(x2 + (w2 - x2) / 2), int(y2 + (h2 - y2) / 2))
                # #     cv2.line(frame, p0, p1, color, 3)
                # #
                # #     if self.intersect(p0, p1, self.line[0], self.line[1]):
                # #         self.counter += 1
                # #
                # # # text = "{}: {:.4f}".format(LABELS[classIDs[i]], confidences[i])
                # # text = "{}".format(indexIDs[i])
                # # cv2.putText(frame, text, (x, y - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)
                # i += 1

                label = '{} {:.2f}'.format(indexIDs[i], out_scores[i])
                draw = ImageDraw.Draw(image)
                label_size = draw.textsize(label, font)

                left, top, right,bottom = box
                top = max(0, np.floor(top + 0.5).astype('int32'))
                left = max(0, np.floor(left + 0.5).astype('int32'))
                bottom = min(image.size[1], np.floor(bottom + 0.5).astype('int32'))
                right = min(image.size[0], np.floor(right + 0.5).astype('int32'))
                # print(label, (left, top), (right, bottom))

                if top - label_size[1] >= 0:
                    text_origin = np.array([left, top - label_size[1]])
                else:
                    text_origin = np.array([left, top + 1])

                # My kingdom for a good redistributable image drawing library.
                for j in range(thickness):
                    draw.rectangle(
                        [left + j, top + j, right - j, bottom - j],
                        outline=(255, 0, 248))
                draw.rectangle(
                    [tuple(text_origin), tuple(text_origin + label_size)],
                    fill=(255, 0, 248))
                draw.text(text_origin, label, fill=(0, 0, 0), font=font)


                if indexIDs[i] in previous:
                    previous_box = previous[indexIDs[i]]
                    (x2, y2) = (int(previous_box[0]), int(previous_box[1]))
                    (w2, h2) = (int(previous_box[2]), int(previous_box[3]))
                    p0 = (int(left + (right - left) / 2), int(top + (bottom - top) / 2))

                    p1 = (int(x2 + (w2 - x2) / 2), int(y2 + (h2 - y2) / 2))
                    # cv2.line(frame, p0, p1, color, 3)
                    draw.line((p0[0],p0[1], p1[0],p1[1]), fill=(255,0, 255), width=3)

                    # L_Vehicle = self.lineFunc([p0[0], p0[1]], [p1[0], p1[1]])
                    #
                    # L_left = self.lineFunc([line_left[0][0], line_left[0][1]], [line_left[1][0], line_left[1][1]])
                    # L_mid = self.lineFunc([lines_mid[0][0], lines_mid[0][1]], [lines_mid[1][0], lines_mid[1][1]])
                    # L_right = self.lineFunc([line_right[0][0], line_right[0][1]], [line_right[1][0], line_right[1][1]])
                    #
                    # if (self.intersection(L_Vehicle, L_left)):
                    #     self.left = True
                    #     self.right=False
                    # #     print("Intersection detected:")
                    # # else:
                    # #     print("No single intersection point detected")
                    #
                    # if(self.intersection(L_Vehicle, L_mid)):
                    #     self.mid=True
                    #     if (self.left and self.mid):
                    #         self.leftToright_counter += 1
                    #         self.mid=False
                    #         self.left=False
                    #         self.right=False
                    #     elif (self.right and self.mid):
                    #         print("r2l- L:{},M:{},R:{}".format(self.left, self.mid, self.right))
                    #         self.rightToleft_counter += 1
                    #         self.mid=False
                    #         self.right=False
                    #         self.left=False
                    #     print("Fin- L:{},M:{},R:{}".format(self.left, self.mid, self.right))
                    #     self.mid = False
                    #     self.right = False
                    #     self.left = False
                    #
                    # if (self.intersection(L_Vehicle, L_right)):
                    #     self.right=True
                    #     self.left=False


                    if self.intersect(p0, p1, line_left[0], line_left[1]):
                        # self.left_counter += 1
                        self.left=True
                        self.right=False

                    if self.intersect(p0, p1, lines_mid[0], lines_mid[1]):
                        self.mid=True

                        if (self.left and self.mid):
                            print("l2r- L:{},M:{},R:{}".format(self.left, self.mid, self.right))
                            self.leftToright_counter += 1
                            self.mid=False
                            self.left=False
                            self.right=False

                        elif (self.right and self.mid):
                            print("r2l- L:{},M:{},R:{}".format(self.left, self.mid, self.right))
                            self.rightToleft_counter += 1
                            self.mid=False
                            self.right=False
                            self.left=False

                        print("Fin- L:{},M:{},R:{}".format(self.left, self.mid, self.right))
                        self.mid = False
                        self.right = False
                        self.left = False

                    if self.intersect(p0, p1, line_right[0], line_right[1]):
                        self.right=True
                        self.left=False



                del draw

                i += 1
        frame=np.asarray(image)
        # draw line
        cv2.line(frame, lines_mid[0], lines_mid[1], (0, 255, 255), 5)
        cv2.line(frame, line_left[0], line_left[1], (255, 255, 255), 5)
        cv2.line(frame, line_right[0], line_right[1], (255, 255, 255), 5)

        # draw counter
        # cv2.putText(frame, str(self.counter), (100, 200), cv2.FONT_HERSHEY_DUPLEX, 5.0, (0, 255, 255), 10)
        cv2.putText(frame, str("L2R:{}".format(self.leftToright_counter)), (350, 200), cv2.FONT_HERSHEY_DUPLEX, 1.0, (255, 0, 255), 2)
        cv2.putText(frame, str("R2L:{}".format(self.rightToleft_counter)), (100, 200), cv2.FONT_HERSHEY_DUPLEX, 1.0, (255, 255, 0), 2)
        # rightToleft_counter
        # counter += 1
        # Display the resulting frame
        # cv2.imshow('Frame', frame)

        image=Image.fromarray(frame)
        return image,self.leftToright_counter,self.rightToleft_counter
        #
        # # for i, c in reversed(list(enumerate(out_classes))):
        # for i in range(len(out_boxes)):
        #     # predicted_class = self.class_names[c]
        #     box = out_boxes[i]
        #     score = out_scores[i]
        #
        #     label = '{} {:.2f}'.format("cl", score)
        #     draw = ImageDraw.Draw(image)
        #     label_size = draw.textsize(label, font)
        #
        #     top, left, bottom, right = box
        #     top = max(0, np.floor(top + 0.5).astype('int32'))
        #     left = max(0, np.floor(left + 0.5).astype('int32'))
        #     bottom = min(image.size[1], np.floor(bottom + 0.5).astype('int32'))
        #     right = min(image.size[0], np.floor(right + 0.5).astype('int32'))
        #     print(label, (left, top), (right, bottom))
        #
        #     if top - label_size[1] >= 0:
        #         text_origin = np.array([left, top - label_size[1]])
        #     else:
        #         text_origin = np.array([left, top + 1])
        #
        #     # My kingdom for a good redistributable image drawing library.
        #     for i in range(thickness):
        #         draw.rectangle(
        #             [left + i, top + i, right - i, bottom - i],
        #             outline=self.colors[1])
        #     draw.rectangle(
        #         [tuple(text_origin), tuple(text_origin + label_size)],
        #         fill=self.colors[1])
        #     draw.text(text_origin, label, fill=(0, 0, 0), font=font)
        #
        #     if indexIDs[i] in previous:
        #         previous_box = previous[indexIDs[i]]
        #         (x2, y2) = (int(previous_box[0]), int(previous_box[1]))
        #         (w2, h2) = (int(previous_box[2]), int(previous_box[3]))
        #         p0 = (int(top + (bottom - top) / 2), int(left + (right - left) / 2))
        #         p1 = (int(x2 + (w2 - x2) / 2), int(y2 + (h2 - y2) / 2))
        #         # cv2.line(frame, p0, p1, color, 3)
        #
        #         if self.intersect(p0, p1, self.line[0], self.line[1]):
        #             self.counter += 1
        #
        #     # text = "{}: {:.4f}".format(LABELS[classIDs[i]], confidences[i])
        #     text = "{}".format(indexIDs[i])
        #     print(text)
        #     del draw
        #
        # end = timer()
        # print(end - start)
        #
        # return image





        # if len(boxes) > 0:
        #     i = int(0)
        #     for box in boxes:
        #         # extract the bounding box coordinates
        #         (x, y) = (int(box[0]), int(box[1]))
        #         (w, h) = (int(box[2]), int(box[3]))
        #
        #         # draw a bounding box rectangle and label on the image
        #         # color = [int(c) for c in COLORS[classIDs[i]]]
        #         # cv2.rectangle(frame, (x, y), (x + w, y + h), color, 2)
        #
        #         label = '{} {:.2f}'.format(predicted_class, score)
        #         # draw = ImageDraw.Draw(image)
        #         # label_size = draw.textsize(label, font)
        #
        #         top, left, bottom, right = box
        #         top = max(0, np.floor(top + 0.5).astype('int32'))
        #         left = max(0, np.floor(left + 0.5).astype('int32'))
        #         bottom = min(image.size[1], np.floor(bottom + 0.5).astype('int32'))
        #         right = min(image.size[0], np.floor(right + 0.5).astype('int32'))
        #         print(label, (left, top), (right, bottom))
        #
        #         if top - label_size[1] >= 0:
        #             text_origin = np.array([left, top - label_size[1]])
        #         else:
        #             text_origin = np.array([left, top + 1])
        #
        #         # My kingdom for a good redistributable image drawing library.
        #         for i in range(thickness):
        #             draw.rectangle(
        #                 [left + i, top + i, right - i, bottom - i],
        #                 outline=self.colors[c])
        #         draw.rectangle(
        #             [tuple(text_origin), tuple(text_origin + label_size)],
        #             fill=self.colors[c])
        #         draw.text(text_origin, label, fill=(0, 0, 0), font=font)
        #     del draw
        #
        #         result = np.asarray(image)
        #         color = [int(c) for c in self.COLORS[indexIDs[i] % len(self.COLORS)]]
        #         cv2.rectangle(result, (x, y), (w, h), (255,0,0), 2)
        #
        #         if indexIDs[i] in previous:
        #             previous_box = previous[indexIDs[i]]
        #             (x2, y2) = (int(previous_box[0]), int(previous_box[1]))
        #             (w2, h2) = (int(previous_box[2]), int(previous_box[3]))
        #             p0 = (int(x + (w - x) / 2), int(y + (h - y) / 2))
        #             p1 = (int(x2 + (w2 - x2) / 2), int(y2 + (h2 - y2) / 2))
        #             cv2.line(result, p0, p1, color, 3)
        #
        #             if self.intersect(p0, p1, self.line[0], self.line[1]):
        #                 self.counter += 1
        #
        #         # text = "{}: {:.4f}".format(LABELS[classIDs[i]], confidences[i])
        #         text = "{}".format(indexIDs[i])
        #         cv2.putText(result, text, (x, y - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)
        #         i += 1
        # # # draw line
        # cv2.line(result, self.line[0], self.line[1], (0, 255, 255), 5)
        # #
        # # # draw counter
        # cv2.putText(result, str(self.counter), (100, 200), cv2.FONT_HERSHEY_DUPLEX, 5.0, (0, 255, 255), 10)
        # self.counter += 1
        # # Display the resulting frame
        # cv2.imshow('Frame', result)

        # for i, c in reversed(list(enumerate(out_classes))):
        #     predicted_class = self.class_names[c]
        #     box = out_boxes[i]
        #     score = out_scores[i]
        #
        #     label = '{} {:.2f}'.format(predicted_class, score)
        #     # draw = ImageDraw.Draw(image)
        #     # label_size = draw.textsize(label, font)
        #
        #     top, left, bottom, right = box
        #     top = max(0, np.floor(top + 0.5).astype('int32'))
        #     left = max(0, np.floor(left + 0.5).astype('int32'))
        #     bottom = min(image.size[1], np.floor(bottom + 0.5).astype('int32'))
        #     right = min(image.size[0], np.floor(right + 0.5).astype('int32'))
        #     print(label, (left, top), (right, bottom))
        #
        #     # if top - label_size[1] >= 0:
        #     #     text_origin = np.array([left, top - label_size[1]])
        #     # else:
        #     #     text_origin = np.array([left, top + 1])
        #
        #     # My kingdom for a good redistributable image drawing library.
        #     # for i in range(thickness):
        #     #     # draw.rectangle(
        #     #     #     [left + i, top + i, right - i, bottom - i],
        #     #         outline=self.colors[c])
        #     # draw.rectangle(
        #     #     [tuple(text_origin), tuple(text_origin + label_size)],
        #     #     fill=self.colors[c])
        #     # draw.text(text_origin, label, fill=(0, 0, 0), font=font)
            # del draw

        # end = timer()
        # print(end - start)
        # return result

    def close_session(self):
        self.sess.close()

# yolo1 = YOLO
def get_points(im):
    # Set up data to send to mouse handler
    data = {}
    data['im'] = im.copy()
    data['lines'] = []

    # Set the callback function for any mouse event
    cv2.imshow("Image", im)
    cv2.setMouseCallback("Image", mouse_handler, data)
    # cv2.waitKey(0)

    # Convert array to np.array in shape n,2,2
    # points = np.uint16(data['lines'])

    return data['im']

def mouse_handler(event, x, y, flags, data):
    global btn_down

    if event == cv2.EVENT_LBUTTONUP and btn_down:
        #if you release the button, finish the line
        btn_down = False
        data['lines'][0].append((x, y)) #append the second point
        # print('up',x,y)
        lines.append((x,y))
        cv2.circle(data['im'], (x, y), 3, (0, 0, 255),5)
        cv2.line(data['im'], data['lines'][0][0], data['lines'][0][1], (0,0,255), 2)
        cv2.imshow("Image", data['im'])

        if len(data['lines']) == numberOfLines:
            cv2.setMouseCallback("Image", lambda *args: None)

            print('lines',lines)
            # showImage(YOLO.lines)
            detect_video(lines)
            cv2.destroyWindow('Image')


    elif event == cv2.EVENT_MOUSEMOVE and btn_down:
        #thi is just for a ine visualization
        image = data['im'].copy()
        cv2.line(image, data['lines'][0][0], (x, y), (0,0,0), 1)
        cv2.imshow("Image", image)

    elif event == cv2.EVENT_LBUTTONDOWN and len(data['lines']) < numberOfLines:

        btn_down = True
        data['lines'].insert(0,[(x, y)]) #prepend the point
        # print('down',x,y)
        lines.append((x, y))
        cv2.circle(data['im'], (x, y), 3, (0, 0, 255), 5, 16)
        cv2.imshow("Image", data['im'])

        # print(data['lines'])

def getFirstFrame(videofile):
    vidcap = cv2.VideoCapture(videofile)
    success, image = vidcap.read()
    if success:
        # cv2.imwrite("first_frame.jpg", image)  # save frame as JPEG file
        vidcap.release()
        return image

def detect_start(yolo1: object, video_path1: object, output_path1: object = "") -> object:
    global yolo,video_path,output_path
    yolo=yolo1
    video_path=video_path1
    output_path=output_path1
    img = getFirstFrame(video_path)
    print('lines', numberOfLines)
    final_image = get_points(img)
    cv2.waitKey(0)
    # cv2.destroyWindow('Image')
    # print('lines', lines)

def write_to_excel(time_count,duration_mins,entry,exit,filepath='./output.xlsx'):

    if (os.path.exists(filepath)):
        workbook_obj = openpyxl.load_workbook(filepath)
        sheet_obj = workbook_obj.active
        start_time = (time_count - 1) * duration_mins
        finish_time = (time_count - 1) * duration_mins + duration_mins
        # '%02d:%02d' % start_time,1
        col1 = '%02d:%02d' % (start_time,1)
        col2 = '%02d:%02d' % (finish_time, 0)
        print(col1)
        print(col2)
        col3 = entry
        col4 = exit
        sheet_obj.append([col1, col2, col3, col4])
    else:
        workbook_obj = openpyxl.Workbook()
        sheet_obj = workbook_obj.active
        col1 = 'Time Start'
        col2 = 'Time End'
        col3 = 'Entry'
        col4 = 'Exit'
        sheet_obj.append([col1, col2, col3, col4])

        start_time = (time_count - 1) * duration_mins
        finish_time = (time_count - 1) * duration_mins + duration_mins

        col1 = '%02d:%02d' % (start_time,1)
        col2 = '%02d:%02d' % (finish_time, 0)
        # print(col1)
        # print(col2)
        col3 = entry
        col4 = exit
        sheet_obj.append([col1, col2, col3, col4])

    workbook_obj.save(filepath)


    # workbook = xlsxwriter.Workbook('output.xlsx')
    #
    # # The workbook object is then used to add new
    # # worksheet via the add_worksheet() method.
    # worksheet = workbook.add_worksheet()
    #
    # # Use the worksheet object to write
    # # data via the write() method.
    # worksheet.write('Time Start', 'Hello..')
    # worksheet.write('Time End', 'Geeks')
    # worksheet.write('Entry', 'For')
    # worksheet.write('Exit', 'Geeks')
    #
    # # Finally, close the Excel file
    # # via the close() method.
    # workbook.close()
# def detect_video(yolo, video_path, output_path=""):
def detect_video(lines_mid):

    global yolo, video_path, output_path

    #declare the duration here
    duration_minutes=5


    time_count=0
    init_l2r=0
    init_r2l=0

    print("line clicked",lines_mid)
    import cv2
    vid = cv2.VideoCapture(video_path)
    if not vid.isOpened():
        raise IOError("Couldn't open webcam or video")
    video_FourCC    = int(vid.get(cv2.CAP_PROP_FOURCC))
    video_fps       = vid.get(cv2.CAP_PROP_FPS)
    frame_count = int(vid.get(cv2.CAP_PROP_FRAME_COUNT))

    video_size      = (int(vid.get(cv2.CAP_PROP_FRAME_WIDTH)),
                        int(vid.get(cv2.CAP_PROP_FRAME_HEIGHT)))
    isOutput = True if output_path != "" else False
    if isOutput:
        print("!!! TYPE:", type(output_path), type(video_FourCC), type(video_fps), type(video_size))
        out = cv2.VideoWriter(output_path, video_FourCC, video_fps, video_size)

    cv2.destroyWindow('Image')
    accum_time = 0
    curr_fps = 0
    fps = "FPS: ??"
    prev_time = timer()
    frame_num=0
    while True:
        return_value, frame = vid.read()
        image = Image.fromarray(frame)


        image,l2r_count,r2l_count = yolo.detect_image(image,lines_mid)
        result = np.asarray(image)


        curr_time = timer()
        exec_time = curr_time - prev_time
        prev_time = curr_time
        accum_time = accum_time + exec_time
        curr_fps = curr_fps + 1
        if accum_time > 1:
            accum_time = accum_time - 1
            fps = "FPS: " + str(curr_fps)
            curr_fps = 0
        cv2.putText(result, text=fps, org=(3, 15), fontFace=cv2.FONT_HERSHEY_SIMPLEX,
                    fontScale=0.50, color=(255, 0, 0), thickness=2)



        cv2.namedWindow("result", cv2.WINDOW_NORMAL)
        cv2.imshow("result", result)

        frame_num += 1
        duration = frame_num / video_fps
        minutes = int(duration / 60)
        seconds = duration % 60

        if minutes%duration_minutes==0 and seconds==0:
            print("Frame:{}, min: {},seconds: {}".format(str(frame_num), str(minutes),str(seconds)))
            time_count+=1
            init_l2r=l2r_count-init_l2r
            init_r2l=r2l_count-init_r2l
            print("L2R:{}, R2L: {}".format(str(init_l2r), str(init_r2l)))

            write_to_excel(time_count=time_count,duration_mins=duration_minutes,entry=init_l2r,exit=init_r2l,filepath='./output.xlsx')

            # write_to_excel(time_count=time_count, duration_mins=duration_minutes, entry=init_r2l, exit=init_l2r,
            #                filepath='./output.xlsx')


        if isOutput:
            out.write(result)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    yolo.close_session()

